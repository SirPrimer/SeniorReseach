import os
import openpyxl


folder_path = './test'


for root, dirs, files in os.walk(folder_path):
    for filename in files:

        if filename.endswith('.xlsx') or filename.endswith('.xls'):

            workbook = openpyxl.load_workbook(os.path.join(root, filename))

            for sheetname in workbook.sheetnames:
                sheet = workbook[sheetname]

                if sheet.min_row == sheet.max_row == 1 and sheet.min_column == sheet.max_column == 1 and sheet.cell(1, 1).value is None:

                    workbook.remove(sheet)

            workbook.save(os.path.join(root, filename))